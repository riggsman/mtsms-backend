"""Parse CSV/XLSX academic calendars and upsert entries by event date."""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import BinaryIO, List, Optional, Tuple

from dateutil import parser as date_parser

MAX_IMPORT_ROWS = 500
MAX_FILE_BYTES = 5 * 1024 * 1024

DATE_RANGE_SEP_RE = re.compile(r"\s*(?:–|—|to)\s*", re.IGNORECASE)

DATE_HEADERS = frozenset({"DATE", "DATES", "EVENT DATE", "DAY"})
ACTIVITY_HEADERS = frozenset({
    "ACTIVITIES",
    "ACTIVITY",
    "EVENT",
    "EVENTS",
    "PROGRAM",
    "PROGRAMS",
    "DESCRIPTION",
})


@dataclass
class ParsedCalendarRow:
    event_date: date
    event_end_date: date
    activity: str
    row_number: int
    row_order: int


def format_date_display(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def format_date_range_display(start: date, end: Optional[date] = None) -> str:
    if not start:
        return ""
    end = end or start
    if end == start:
        return format_date_display(start)
    return f"{format_date_display(start)} – {format_date_display(end)}"


@dataclass
class ParseError:
    row: int
    message: str


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    return re.sub(r"\s+", " ", text)


def _map_headers(headers: List[str]) -> Tuple[Optional[int], Optional[int]]:
    date_idx = None
    activity_idx = None
    for i, raw in enumerate(headers):
        norm = _normalize_header(raw)
        if norm in DATE_HEADERS:
            date_idx = i
        elif norm in ACTIVITY_HEADERS:
            activity_idx = i
    if date_idx is None and activity_idx is None and len(headers) == 2:
        return 0, 1
    return date_idx, activity_idx


def _parse_date_value(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(value).date()
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        parsed = date_parser.parse(text, dayfirst=True)
        return parsed.date()
    except (ValueError, TypeError, OverflowError):
        return None


def _parse_date_range_value(value) -> Optional[Tuple[date, date]]:
    if value is None:
        return None
    if isinstance(value, datetime):
        d = value.date()
        return d, d
    if isinstance(value, date):
        return value, value
    if isinstance(value, (int, float)):
        d = _parse_date_value(value)
        return (d, d) if d else None

    text = str(value).strip()
    if not text:
        return None

    if DATE_RANGE_SEP_RE.search(text):
        parts = DATE_RANGE_SEP_RE.split(text, maxsplit=1)
        if len(parts) == 2:
            start = _parse_date_value(parts[0].strip())
            end = _parse_date_value(parts[1].strip())
            if start and end:
                if end < start:
                    start, end = end, start
                return start, end
            return None

    single = _parse_date_value(text)
    return (single, single) if single else None


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_blank_row(cells: List) -> bool:
    return all(_cell_str(c) == "" for c in cells)


def parse_csv_content(content: bytes) -> Tuple[List[ParsedCalendarRow], List[ParseError]]:
    if len(content) > MAX_FILE_BYTES:
        return [], [ParseError(row=0, message=f"File exceeds {MAX_FILE_BYTES // (1024 * 1024)} MB limit")]
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("latin-1")
        except UnicodeDecodeError:
            return [], [ParseError(row=0, message="Unable to decode CSV file")]

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    return _parse_tabular_rows(rows)


def parse_xlsx_content(content: bytes) -> Tuple[List[ParsedCalendarRow], List[ParseError]]:
    if len(content) > MAX_FILE_BYTES:
        return [], [ParseError(row=0, message=f"File exceeds {MAX_FILE_BYTES // (1024 * 1024)} MB limit")]
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], [ParseError(row=0, message="openpyxl is not installed")]

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    tabular: List[List] = []
    for row in ws.iter_rows(values_only=True):
        tabular.append(list(row))
    wb.close()
    return _parse_tabular_rows(tabular)


def parse_upload_file(filename: str, content: bytes) -> Tuple[List[ParsedCalendarRow], List[ParseError]]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return parse_csv_content(content)
    if lower.endswith(".xlsx"):
        return parse_xlsx_content(content)
    return [], [ParseError(row=0, message="Only .csv and .xlsx files are supported")]


def _parse_tabular_rows(rows: List[List]) -> Tuple[List[ParsedCalendarRow], List[ParseError]]:
    errors: List[ParseError] = []
    parsed: List[ParsedCalendarRow] = []
    if not rows:
        return [], [ParseError(row=0, message="File is empty")]

    header_row = rows[0]
    date_idx, activity_idx = _map_headers([_cell_str(c) for c in header_row])
    if date_idx is None or activity_idx is None:
        return [], [
            ParseError(
                row=1,
                message="Missing required columns. Use DATE and ACTIVITIES (or two columns without headers).",
            )
        ]

    data_rows = rows[1:]
    non_blank = 0
    order = 0
    for offset, raw in enumerate(data_rows, start=2):
        cells = list(raw) if raw else []
        if _is_blank_row(cells):
            continue
        non_blank += 1
        if non_blank > MAX_IMPORT_ROWS:
            errors.append(ParseError(row=offset, message=f"Maximum {MAX_IMPORT_ROWS} data rows exceeded"))
            break

        date_val = cells[date_idx] if date_idx < len(cells) else None
        activity_val = cells[activity_idx] if activity_idx < len(cells) else None
        date_range = _parse_date_range_value(date_val)
        activity = _cell_str(activity_val)

        if date_range is None:
            errors.append(
                ParseError(
                    row=offset,
                    message=f"Invalid date or range: {_cell_str(date_val) or '(empty)'}",
                )
            )
            continue
        event_date, event_end_date = date_range
        if not activity:
            errors.append(ParseError(row=offset, message="Activity is required"))
            continue

        order += 1
        parsed.append(
            ParsedCalendarRow(
                event_date=event_date,
                event_end_date=event_end_date,
                activity=activity,
                row_number=offset,
                row_order=order,
            )
        )

    return parsed, errors


def upsert_calendar_rows(
    db,
    *,
    institution_id: int,
    academic_year_id: int,
    rows: List[ParsedCalendarRow],
    source_filename: Optional[str],
    uploaded_by_user_id: Optional[int],
) -> Tuple[int, int, int]:
    """Returns (imported, updated, skipped)."""
    from app.models.academic_calendar import AcademicCalendar

    imported = 0
    updated = 0
    skipped = 0
    now = datetime.utcnow()

    existing = (
        db.query(AcademicCalendar)
        .filter(
            AcademicCalendar.institution_id == institution_id,
            AcademicCalendar.academic_year_id == academic_year_id,
            AcademicCalendar.deleted_at.is_(None),
        )
        .all()
    )
    by_range = {(r.event_date, r.event_end_date or r.event_date): r for r in existing}

    for row in rows:
        key = (row.event_date, row.event_end_date)
        current = by_range.get(key)
        if current:
            if (
                current.activity == row.activity
                and (current.row_order or 0) == row.row_order
            ):
                skipped += 1
                continue
            current.activity = row.activity
            current.row_order = row.row_order
            current.source_filename = source_filename
            current.uploaded_by_user_id = uploaded_by_user_id
            current.updated_at = now
            updated += 1
        else:
            entry = AcademicCalendar(
                institution_id=institution_id,
                academic_year_id=academic_year_id,
                event_date=row.event_date,
                event_end_date=row.event_end_date,
                activity=row.activity,
                row_order=row.row_order,
                source_filename=source_filename,
                uploaded_by_user_id=uploaded_by_user_id,
                created_at=now,
                updated_at=now,
            )
            db.add(entry)
            by_range[key] = entry
            imported += 1

    db.commit()
    return imported, updated, skipped


TEMPLATE_CSV = (
    "DATE,ACTIVITIES\n"
    "01/09/2026 – 30/09/2026,Semester opening period\n"
    "15/12/2026,End of semester exams\n"
)
